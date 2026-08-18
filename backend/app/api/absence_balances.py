import io
from datetime import date, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_impersonation_employee
from app.db import get_db
from app.models import Employee, EmployeeAbsenceBalance, EmployeeAbsenceBalanceStatus, User
from app.schemas import (
    EmployeeAbsenceBalanceRead,
    EmployeeAbsenceBalancesCommit,
    EmployeeAbsenceBalancesCommitRead,
    EmployeeAbsenceBalanceStatusRead,
    EmployeeAbsenceBalanceUpdate,
)
from app.services.audit import record_audit_log
from app.services.portal_auth import build_auth_user_read, build_impersonation_view
from app.services.security import get_current_user


router = APIRouter(prefix="/absence-balances", tags=["absence-balances"])
ZERO = Decimal("0.00")
STATUS_ROW_ID = 1


def _decimal_string(value: Decimal) -> str:
    return format(value, ".2f")


def _effective_auth(db: Session, current_user: User, impersonate_employee: Employee | None):
    if impersonate_employee is not None:
        return build_impersonation_view(db, impersonate_employee)
    return build_auth_user_read(db, current_user)


def _require_admin_or_hr(auth) -> None:
    if auth.effective_role not in {"admin", "hr"} or not auth.can_access_calendar:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Accesso riservato ad admin e HR.")


def _serialize(employee: Employee, balance: EmployeeAbsenceBalance | None) -> EmployeeAbsenceBalanceRead:
    return EmployeeAbsenceBalanceRead(
        employee_id=employee.id,
        employee_name=employee.full_name,
        tms_id=employee.tms_id,
        permission_hours=float(balance.permission_hours if balance else ZERO),
        vacation_days=float(balance.vacation_days if balance else ZERO),
        last_modified_at=balance.updated_at if balance else None,
        last_modified_by=balance.updated_by_name if balance else None,
    )


def _list_rows(db: Session) -> list[EmployeeAbsenceBalanceRead]:
    rows = db.execute(
        select(Employee, EmployeeAbsenceBalance)
        .outerjoin(EmployeeAbsenceBalance, EmployeeAbsenceBalance.employee_id == Employee.id)
        .where(Employee.is_active.is_(True))
        .order_by(Employee.full_name.asc())
    ).all()
    return [_serialize(employee, balance) for employee, balance in rows]


def _previous_month_end(today: date | None = None) -> date:
    current = today or date.today()
    return date(current.year, current.month, 1) - timedelta(days=1)


def _serialize_status(status_row: EmployeeAbsenceBalanceStatus | None) -> EmployeeAbsenceBalanceStatusRead:
    return EmployeeAbsenceBalanceStatusRead(
        updated_through=status_row.updated_through if status_row else None,
        last_modified_at=status_row.updated_at if status_row else None,
        last_modified_by=status_row.updated_by_name if status_row else None,
    )


def _update_balance_row(
    db: Session,
    *,
    employee: Employee,
    payload: EmployeeAbsenceBalanceUpdate,
    current_user: User,
) -> EmployeeAbsenceBalance:
    balance = db.get(EmployeeAbsenceBalance, employee.id)
    before = {
        "permission_hours": _decimal_string(balance.permission_hours if balance else ZERO),
        "vacation_days": _decimal_string(balance.vacation_days if balance else ZERO),
    }
    if balance is None:
        balance = EmployeeAbsenceBalance(employee_id=employee.id)
        db.add(balance)
    balance.permission_hours = payload.permission_hours
    balance.vacation_days = payload.vacation_days
    balance.updated_by_user_id = current_user.id
    balance.updated_by_name = current_user.display_name or current_user.username
    db.flush()

    after = {
        "permission_hours": _decimal_string(balance.permission_hours),
        "vacation_days": _decimal_string(balance.vacation_days),
    }
    record_audit_log(
        db,
        action="update",
        entity="employee_absence_balance",
        actor_name=current_user.username,
        user_id=current_user.id,
        detail={"employee_id": employee.id, "employee_name": employee.full_name, "before": before, "after": after},
    )
    return balance


@router.get("", response_model=list[EmployeeAbsenceBalanceRead])
def list_absence_balances(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> list[EmployeeAbsenceBalanceRead]:
    auth = _effective_auth(db, current_user, impersonate_employee)
    _require_admin_or_hr(auth)
    return _list_rows(db)


@router.get("/status", response_model=EmployeeAbsenceBalanceStatusRead)
def get_absence_balance_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> EmployeeAbsenceBalanceStatusRead:
    auth = _effective_auth(db, current_user, impersonate_employee)
    _require_admin_or_hr(auth)
    return _serialize_status(db.get(EmployeeAbsenceBalanceStatus, STATUS_ROW_ID))


@router.post("/commit", response_model=EmployeeAbsenceBalancesCommitRead)
def commit_absence_balances(
    payload: EmployeeAbsenceBalancesCommit,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> EmployeeAbsenceBalancesCommitRead:
    auth = _effective_auth(db, current_user, impersonate_employee)
    _require_admin_or_hr(auth)
    if not auth.can_edit_absence_balances:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non sei abilitato a modificare ferie e permessi residui.",
        )

    expected_date = _previous_month_end()
    if payload.updated_through != expected_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"La data di aggiornamento deve essere {expected_date.strftime('%d/%m/%Y')}.",
        )

    employee_ids = [change.employee_id for change in payload.changes]
    if len(employee_ids) != len(set(employee_ids)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ogni dipendente può comparire una sola volta.")

    employees = {
        employee.id: employee
        for employee in db.scalars(
            select(Employee).where(Employee.id.in_(employee_ids), Employee.is_active.is_(True))
        ).all()
    }
    missing_ids = [employee_id for employee_id in employee_ids if employee_id not in employees]
    if missing_ids:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uno o più dipendenti non sono stati trovati.")

    saved: list[tuple[Employee, EmployeeAbsenceBalance]] = []
    for change in payload.changes:
        employee = employees[change.employee_id]
        balance = _update_balance_row(db, employee=employee, payload=change, current_user=current_user)
        saved.append((employee, balance))

    status_row = db.get(EmployeeAbsenceBalanceStatus, STATUS_ROW_ID)
    previous_date = status_row.updated_through if status_row else None
    if status_row is None:
        status_row = EmployeeAbsenceBalanceStatus(id=STATUS_ROW_ID)
        db.add(status_row)
    status_row.updated_through = payload.updated_through
    status_row.updated_by_user_id = current_user.id
    status_row.updated_by_name = current_user.display_name or current_user.username
    record_audit_log(
        db,
        action="update",
        entity="employee_absence_balance_status",
        actor_name=current_user.username,
        user_id=current_user.id,
        detail={
            "before": previous_date.isoformat() if previous_date else None,
            "after": payload.updated_through.isoformat(),
            "updated_employee_ids": employee_ids,
        },
    )
    db.commit()
    for _, balance in saved:
        db.refresh(balance)
    return EmployeeAbsenceBalancesCommitRead(
        updated_through=payload.updated_through,
        balances=[_serialize(employee, balance) for employee, balance in saved],
    )


@router.get("/export")
def export_absence_balances(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> StreamingResponse:
    auth = _effective_auth(db, current_user, impersonate_employee)
    _require_admin_or_hr(auth)
    rows = _list_rows(db)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Residui"
    headers = ["Dipendente", "Matricola", "Ferie (GG)", "Permessi (Ore)", "Ultima Modifica", "Utente Ultima Modifica"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="007040")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append(
            [
                row.employee_name,
                row.tms_id,
                float(row.vacation_days),
                float(row.permission_hours),
                row.last_modified_at.replace(tzinfo=None) if row.last_modified_at else None,
                row.last_modified_by or "",
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max(sheet.max_row, 1)}"
    for column, width in {"A": 32, "B": 14, "C": 14, "D": 18, "E": 22, "F": 28}.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet["E"][1:]:
        cell.number_format = "dd/mm/yyyy hh:mm"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="residui-assenze.xlsx"'},
    )


@router.get("/{employee_id}", response_model=EmployeeAbsenceBalanceRead)
def get_absence_balance(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> EmployeeAbsenceBalanceRead:
    auth = _effective_auth(db, current_user, impersonate_employee)
    can_read_all = auth.effective_role in {"admin", "hr"} and auth.can_access_calendar
    if not can_read_all and auth.linked_employee_id != employee_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Puoi consultare solo i tuoi residui.")
    employee = db.get(Employee, employee_id)
    if employee is None or not employee.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dipendente non trovato.")
    return _serialize(employee, db.get(EmployeeAbsenceBalance, employee_id))


@router.put("/{employee_id}", response_model=EmployeeAbsenceBalanceRead)
def update_absence_balance(
    employee_id: str,
    payload: EmployeeAbsenceBalanceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    impersonate_employee: Employee | None = Depends(get_impersonation_employee),
) -> EmployeeAbsenceBalanceRead:
    auth = _effective_auth(db, current_user, impersonate_employee)
    _require_admin_or_hr(auth)
    if not auth.can_edit_absence_balances:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non sei abilitato a modificare ferie e permessi residui.",
        )

    employee = db.get(Employee, employee_id)
    if employee is None or not employee.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dipendente non trovato.")

    balance = _update_balance_row(db, employee=employee, payload=payload, current_user=current_user)
    db.commit()
    db.refresh(balance)
    return _serialize(employee, balance)
