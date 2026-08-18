import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas import DeviceDeliveryRead


HEADERS = [
    "ID",
    "Stato",
    "Dipendente",
    "Ruolo",
    "Tipo dispositivo",
    "Dispositivo",
    "Numero seriale",
    "Consegnato da",
    "Consegnato il",
    "Restituito il",
    "Note",
]


def export_device_deliveries_xlsx(deliveries: Iterable[DeviceDeliveryRead]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Consegne dispositivi"
    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for delivery in deliveries:
        worksheet.append(
            [
                delivery.id,
                delivery.status,
                delivery.employee_name,
                delivery.employee_role or "",
                delivery.device_asset_type,
                delivery.device_label,
                delivery.device_serial_number or "",
                delivery.delivered_by or "",
                delivery.delivered_at.strftime("%d/%m/%Y %H:%M"),
                delivery.returned_at.strftime("%d/%m/%Y %H:%M") if delivery.returned_at else "",
                delivery.notes or "",
            ]
        )

    for index, _header in enumerate(HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 22
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
