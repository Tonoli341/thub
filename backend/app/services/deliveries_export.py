import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas import EquipmentDeliveryRead


HEADERS = [
    "ID",
    "Dipendente",
    "Ruolo",
    "Categoria",
    "Articolo",
    "Taglia",
    "Quantita",
    "Consegnato da",
    "Consegnato il",
    "Restituito il",
    "Note",
]


def export_deliveries_xlsx(deliveries: Iterable[EquipmentDeliveryRead]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Consegne"
    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for delivery in deliveries:
        worksheet.append(
            [
                delivery.id,
                delivery.employee_name,
                delivery.employee_role or "",
                delivery.item_category,
                delivery.item_name,
                delivery.item_size or "",
                delivery.quantity,
                delivery.delivered_by or "",
                delivery.delivered_at.strftime("%d/%m/%Y %H:%M"),
                delivery.returned_at.strftime("%d/%m/%Y %H:%M") if delivery.returned_at else "",
                delivery.notes or "",
            ]
        )

    for index, _header in enumerate(HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 22
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
