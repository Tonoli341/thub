"""Esportazione Excel per il modulo Manutenzioni (§14 del documento requisiti:
"funzionalità di esportazione generale, in Excel e PDF"). Qui solo la parte
Excel per asset e scadenze: è la richiesta di autonomia di base, non i report
predefiniti per classe che il documento lascia esplicitamente a un secondo
momento.
"""

import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.maintenance_asset_models import MaintenanceAssetCounter, MaintenanceDeadline
from app.maintenance_asset_schemas import MaintenanceAssetRead
from app.services.maintenance_deadlines import compute_urgency

ASSET_HEADERS = [
    "Codice interno",
    "Classe",
    "Sottoclasse",
    "Produttore",
    "Modello",
    "Numero di serie",
    "Sito",
    "Reparto",
    "Responsabile",
    "Stato",
    "Motivo stato",
    "Ultima modifica",
]


def export_maintenance_assets_xlsx(assets: Iterable[MaintenanceAssetRead]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Asset"
    worksheet.append(ASSET_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for asset in assets:
        worksheet.append(
            [
                asset.internal_code,
                asset.asset_class_label,
                asset.asset_type_label,
                asset.custom_fields.get("brand") or "",
                asset.custom_fields.get("model") or "",
                asset.custom_fields.get("serial_number") or "",
                asset.custom_fields.get("site") or "",
                asset.custom_fields.get("department") or "",
                asset.employee_field_names.get("responsible_employee_id") or "",
                asset.status.value,
                asset.status_reason or "",
                asset.updated_at.strftime("%d/%m/%Y %H:%M"),
            ]
        )

    for index, _header in enumerate(ASSET_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 20
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


COUNTER_HEADERS = [
    "Asset",
    "Data lettura",
    "Valore",
    "Unità",
    "Registrato da",
]


def export_maintenance_asset_counters_xlsx(readings: Iterable[MaintenanceAssetCounter]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ore"
    worksheet.append(COUNTER_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for reading in readings:
        worksheet.append(
            [
                reading.asset.internal_code,
                reading.reading_date.strftime("%d/%m/%Y"),
                float(reading.value),
                reading.unit,
                reading.recorded_by or "",
            ]
        )

    for index, _header in enumerate(COUNTER_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 20
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


DEADLINE_HEADERS = [
    "Asset",
    "Tipo scadenza",
    "Scadenza",
    "Stato",
    "Ultimo completamento",
    "Motivo posticipo",
]

URGENCY_LABELS = {"regolare": "Regolare", "in_scadenza": "In scadenza", "urgente": "Urgente", "scaduta": "Scaduta"}


def export_maintenance_deadlines_xlsx(db: Session, deadlines: Iterable[MaintenanceDeadline]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Scadenze"
    worksheet.append(DEADLINE_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for deadline in deadlines:
        worksheet.append(
            [
                deadline.asset.internal_code,
                deadline.deadline_type,
                deadline.due_date.strftime("%d/%m/%Y"),
                URGENCY_LABELS[compute_urgency(db, deadline)],
                deadline.last_completed_at.strftime("%d/%m/%Y") if deadline.last_completed_at else "",
                deadline.postponed_reason or "",
            ]
        )

    for index, _header in enumerate(DEADLINE_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 22
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
